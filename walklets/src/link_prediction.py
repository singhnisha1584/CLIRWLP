import numpy as np
import networkx as nx
import pandas as pd
import random
from sklearn.linear_model import LogisticRegressionCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import roc_auc_score
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegressionCV
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import train_test_split
from sklearn.metrics import *
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC


def gen_rand_edges(teCt, unique_nodes):
	num = 3*len(teCt)
	seen = set(teCt)
	x, y = random.choice(unique_nodes), random.choice(unique_nodes)
	t = 0
	while t < num:
		seen.add((x, y))
		t = t + 1
		x, y = random.choice(unique_nodes), random.choice(unique_nodes)
		while (x, y) in seen or (y, x) in seen:
			x, y = random.choice(unique_nodes), random.choice(unique_nodes)
	
	seen = list(seen)
	return seen


def eval_report(model,x_train, y_train, x_test, y_test):
	test_pred = model.predict(x_test)
	test_acc = accuracy_score(y_test, test_pred)
	prec_per, recall_per, threshold_per = precision_recall_curve(y_test, test_pred)
	prec_per = prec_per[::-1]
	recall_per = recall_per[::-1]
	aupr_value = np.trapz(prec_per, x=recall_per)
	avg_prec_value = average_precision_score(y_test, test_pred)
	AUC = roc_auc_score(y_test, test_pred)
	test_pred_label = np.copy(test_pred)
	a = np.mean(test_pred_label)

	for i in range(len(test_pred)):
		if test_pred[i] < a:
			test_pred_label[i] = 0
		else:
			test_pred_label[i] = 1
	acc_score_value = accuracy_score(y_test, test_pred_label)
	bal_acc_score_value = balanced_accuracy_score(y_test, test_pred_label)
	f1_value = f1_score(y_test, test_pred_label)
	Recall = recall_score(y_test, test_pred_label)
	Precision = precision_score(y_test, test_pred_label)

	print('\nTest accuracy:' + str(test_acc) + "\nAUC:" + str(AUC) + "\nPrecision:" + str(Precision) +
"\nRecall:" + str(Recall) + "\nAUPR:" + str(aupr_value) + "\nAvgPrecision" + str(avg_prec_value) +
"\nAccScore:" + str(acc_score_value) + "\nBalAccScore:" + str(bal_acc_score_value) + "\nF1:" + str(f1_value))
	return [test_acc, AUC, Precision, Recall, aupr_value, avg_prec_value, acc_score_value, bal_acc_score_value, f1_value]



def evaluate_link_prediction(embeddings : list, nx_G, start_row):
	max_iter = 2000
	
	embeddings.insert(0,[])
 	# Dimensions
	num_rows = len(embeddings)  # Number of rows
	num_columns = len(embeddings[0]) if len(embeddings) else 0  # Number of columns (assuming non-empty)

	print("Number of rows:", num_rows)
	print("Number of columns:", num_columns)
	#Create Training Data
	unique_nodes = list(nx_G.nodes())
	#all_possible_edges = [(x,y) for (x,y) in product(unique_nodes, unique_nodes)]
	#Genrate random false edges instead of product
	all_possible_edges = gen_rand_edges(nx_G.edges(), unique_nodes)
	print(all_possible_edges)
	
	# generate edge features for all pairs of nodes
	edge_features = [(embeddings[int(i)] + embeddings[int(j)]) for i,j in all_possible_edges ]
	print(len(edge_features))
	
	# get current edges in the network
	edges = list(nx_G.edges())

	# create target list, 1 if the pair exists in the network, 0 otherwise
	is_con = [1 if e in edges else 0 for e in all_possible_edges]

	print(sum(is_con))
	X = np.array(edge_features)
	y = is_con

	print(len(X))
	print(len(y))

	# train test split
# train test split
	x_train, x_test, y_train, y_test = train_test_split(X,	y,	test_size = 0.3	)
	lr_clf = LogisticRegressionCV(Cs=10, cv=10, scoring="roc_auc", max_iter=max_iter)
	knn_clf = KNeighborsClassifier(n_neighbors=5, weights='uniform', algorithm='auto', leaf_size=30, p=2, metric='minkowski', metric_params=None, n_jobs=None)
	lda_clf = LinearDiscriminantAnalysis(solver='svd', shrinkage=None, priors=None, n_components=None, store_covariance=False, tol=0.0001)
	gnb_clf = GaussianNB(priors=None, var_smoothing=1e-09)

	rf_clf = RandomForestClassifier(n_estimators=100, criterion='gini', max_depth=None, min_samples_split=2, min_samples_leaf=1, 
				 min_weight_fraction_leaf=0.0, max_leaf_nodes=None, min_impurity_decrease=0.0, bootstrap=True, oob_score=False, 
				 n_jobs=None, random_state=None, verbose=0, warm_start=False, class_weight=None, ccp_alpha=0.0, max_samples=None)
         
	sv_clf = SVC(C=1.0, kernel='rbf', degree=3, gamma='scale', coef0=0.0, shrinking=True, probability=True, tol=0.001, cache_size=200, 
	      class_weight=None, verbose=False, max_iter= -1, decision_function_shape='ovr', break_ties=False, random_state=None)
        
	gbt_clf = GradientBoostingClassifier(loss='log_loss', learning_rate=0.1, n_estimators=100, subsample=1.0, 
				      criterion='friedman_mse', min_samples_split=2, min_samples_leaf=1, min_weight_fraction_leaf=0.0, max_depth=5, 
					  min_impurity_decrease=0.0, init=None, random_state=None, max_features=None, verbose=0, 
					  max_leaf_nodes=None, warm_start=False, validation_fraction=0.1, n_iter_no_change=None, tol=0.0001, ccp_alpha=0.0)
  
	result_list = []
	clf_list = [lr_clf, knn_clf, lda_clf, gnb_clf, rf_clf, sv_clf, gbt_clf]
	clf_name_list = [
		"lr_clf",
		"knn_clf",
		"lda_clf",
		"gnb_clf",
		"rf_clf",
		"sv_clf",
		"gbt_clf",
	]

	excel_path = '/content/drive/MyDrive/_mtp/walklets/results/result.xlsx'

	# excel_data = pd.read_excel(excel_path)
	wb = load_workbook(excel_path)
	sheet = wb.active	
 
	result_params = ["MODELS","test_acc", "AUC", "Precision", "Recall", "aupr_value", "avg_prec_value", "acc_score_value", "bal_acc_score_value", "f1_value"]

	for j, param in enumerate(result_params) :
		cell = sheet.cell(row = start_row, column = j + 1)
		cell.value = param

	for i, clf in enumerate(clf_list):
		# clf = Pipeline(steps=[("sc", StandardScaler()), ("clf", clf)])
		clf.fit(x_train, y_train)
		# y_pred = clf.predict(x_test)
		# y_true = y_test
		print("Results for ", clf_name_list[i], " :")
  		
		data = eval_report(clf, x_train, y_train, x_test, y_test)
		data.insert(0,clf_name_list[i])
  
		for j, value in enumerate(data):
			cell = sheet.cell(row= start_row + i + 1, column = j + 1)
			cell.value = value
		wb.save(excel_path)
	
	return
